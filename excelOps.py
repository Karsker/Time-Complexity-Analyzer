import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart
import os
import numpy as np
import matplotlib.pyplot as plt
def writeRunTimesToSheet(path, inputSizes, runTimes, myrow, mycol, fileName, graphRow):
    wb = load_workbook(path)
    sheet = wb.active
    headRow = myrow
    detailRow = myrow+1
    myrow += 2
    orig_row = myrow

    startCol = mycol
    y_vals = np.array(runTimes)
    sheet.cell(row=headRow, column=mycol).value = os.path.basename(fileName).split('/')[-1] 
    sheet.cell(row=detailRow, column=mycol).value = 'Input Size'
    sheet.cell(row=detailRow, column=mycol+1).value = 'Run Time (secs)' 
    myrow = orig_row
    for size in inputSizes:
        sheet.cell(row=myrow, column = mycol).value = size
        myrow += 1
    myrow = orig_row
    mycol += 1
    for run_time in runTimes:
        sheet.cell(row=myrow, column=mycol).value = run_time
        myrow += 1

    values = Reference(sheet, min_row=orig_row, max_row=myrow-1, min_col=startCol+1, max_col=startCol+1)
    x_values = Reference(sheet, min_col=startCol, max_col=startCol, min_row=orig_row, max_row=myrow-1)
    chart = LineChart()
    chart.add_data(values)
    chart.set_categories(x_values)
    chart.title = os.path.basename(fileName).split('/')[-1]
    chart.x_axis.title = 'Input Size'
    chart.y_axis.title = 'Run Time'
    sheet.add_chart(chart, 'A' + str(graphRow))
    wb.save(path)

